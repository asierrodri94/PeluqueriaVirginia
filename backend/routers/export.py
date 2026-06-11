from fastapi import APIRouter
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
import calendar
from datetime import date
from database import get_conn, tabla_factura, mes_a_db


router = APIRouter(prefix="/api/export", tags=["export"])

DIAS_ES = ["LUNES", "MARTES", "MIÉRCOLES", "JUEVES", "VIERNES", "SÁBADO", "DOMINGO"]
MESES_ES = ["", "ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO",
            "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE"]
TRIMESTRES = ["", "1er TRIMESTRE", "2º TRIMESTRE", "3er TRIMESTRE", "4º TRIMESTRE"]

# Cada mes ocupa 4 columnas (FECHA, DÍA, TOTAL, PERSONAS) + 1 de separación
COLS_POR_MES = 4
SEPARACION = 1


def _trimestre_de_mes(mes: int) -> int:
    return (mes - 1) // 3 + 1


def _meses_del_trimestre(mes: int):
    t = _trimestre_de_mes(mes)
    inicio = (t - 1) * 3 + 1
    return [inicio, inicio + 1, inicio + 2]


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _border() -> Border:
    thin = Side(style="thin")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _col_inicio(idx_mes: int) -> int:
    """Columna de inicio (1-based) para el mes idx_mes (0, 1, 2)."""
    return idx_mes * (COLS_POR_MES + SEPARACION) + 1


def _escribir_mes_trimestre(ws, m, anyo, datos, c0, col3_header, border):
    """Escribe un mes en el bloque de columnas c0..c0+3 del worksheet.
    datos[m][dia] = (count, total)
    """
    num_days = calendar.monthrange(anyo, m)[1]
    total_mes = 0.0
    count_mes = 0

    ws.merge_cells(start_row=1, start_column=c0, end_row=1, end_column=c0 + COLS_POR_MES - 1)
    cell = ws.cell(row=1, column=c0, value=f"{MESES_ES[m]} {anyo}")
    cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=12)
    cell.fill = _fill("4472C4")
    cell.alignment = Alignment(horizontal="center")
    cell.border = border
    for col in range(c0 + 1, c0 + COLS_POR_MES):
        ws.cell(row=1, column=col).fill = _fill("4472C4")
        ws.cell(row=1, column=col).border = border

    for i, h in enumerate(["FECHA", "DÍA", col3_header, "TOTAL (€)"]):
        cell = ws.cell(row=2, column=c0 + i, value=h)
        cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cell.fill = _fill("7094C4")
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    for day in range(1, num_days + 1):
        row = day + 2
        d = date(anyo, m, day)
        weekday = d.weekday()
        is_weekend = weekday >= 5
        day_data = datos[m].get(day)
        count_dia = day_data[0] if day_data else None
        total_dia = day_data[1] if day_data else None
        if total_dia:
            total_mes += total_dia
        if count_dia:
            count_mes += count_dia

        values = [d.strftime("%d/%m/%Y"), DIAS_ES[weekday], count_dia, total_dia]
        for i, val in enumerate(values):
            cell = ws.cell(row=row, column=c0 + i, value=val)
            cell.border = border
            cell.alignment = Alignment(horizontal="left" if i == 0 else "center")
            cell.font = Font(bold=True, color="FF0000", name="Arial", size=10) if is_weekend else Font(name="Arial", size=10)
            if i == 3 and val is not None:
                cell.number_format = '#,##0.00 "€"'

    total_row = num_days + 3
    for i in range(COLS_POR_MES):
        ws.cell(row=total_row, column=c0 + i).border = border
        ws.cell(row=total_row, column=c0 + i).fill = _fill("E2EFDA")
    ws.cell(row=total_row, column=c0).value = "TOTAL"
    ws.cell(row=total_row, column=c0).font = Font(bold=True, name="Arial", size=10)
    pc = ws.cell(row=total_row, column=c0 + 2, value=count_mes)
    pc.font = Font(bold=True, name="Arial", size=10)
    pc.alignment = Alignment(horizontal="center")
    tc = ws.cell(row=total_row, column=c0 + 3, value=round(total_mes, 2))
    tc.font = Font(bold=True, name="Arial", size=10)
    tc.number_format = '#,##0.00 "€"'


@router.get("/trimestre")
def exportar_trimestre(mes: int, anyo: int, seccion: str = "A"):
    meses = _meses_del_trimestre(mes)
    trimestre = _trimestre_de_mes(mes)
    tf = tabla_factura(seccion)
    border = _border()

    datos = {}
    with get_conn() as conn:
        for m in meses:
            mes_db = mes_a_db(m)
            rows = conn.execute(
                f"SELECT dia, COUNT(*) as personas, SUM(importe) as total "
                f"FROM {tf} WHERE mes=? AND anyo=? GROUP BY dia ORDER BY dia",
                (mes_db, anyo),
            ).fetchall()
            datos[m] = {r["dia"]: (r["personas"], round(r["total"], 2)) for r in rows}

    wb = Workbook()
    ws = wb.active
    ws.title = f"{TRIMESTRES[trimestre]} {anyo}"

    col_widths = [14, 12, 10, 13]
    for idx in range(3):
        c0 = _col_inicio(idx)
        for i, w in enumerate(col_widths):
            ws.column_dimensions[get_column_letter(c0 + i)].width = w
        if idx < 2:
            ws.column_dimensions[get_column_letter(c0 + COLS_POR_MES)].width = 2

    for idx, m in enumerate(meses):
        _escribir_mes_trimestre(ws, m, anyo, datos, _col_inicio(idx), "PERS.", border)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    nombre_trim = TRIMESTRES[trimestre].replace(" ", "_")
    filename = f"{nombre_trim}_{anyo}_{'B' if seccion == 'B' else 'A'}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/trimestre/compras")
def exportar_trimestre_compras(mes: int, anyo: int):
    meses = _meses_del_trimestre(mes)
    trimestre = _trimestre_de_mes(mes)
    border = _border()

    datos = {}
    with get_conn() as conn:
        for m in meses:
            mes_db = mes_a_db(m)
            rows = conn.execute(
                "SELECT dia, COUNT(*) as num_compras, SUM(importe) as total "
                "FROM compra WHERE mes=? AND anyo=? GROUP BY dia ORDER BY dia",
                (mes_db, anyo),
            ).fetchall()
            datos[m] = {r["dia"]: (r["num_compras"], round(r["total"], 2)) for r in rows}

    wb = Workbook()
    ws = wb.active
    ws.title = f"{TRIMESTRES[trimestre]} {anyo} Compras"

    col_widths = [14, 12, 13, 13]
    for idx in range(3):
        c0 = _col_inicio(idx)
        for i, w in enumerate(col_widths):
            ws.column_dimensions[get_column_letter(c0 + i)].width = w
        if idx < 2:
            ws.column_dimensions[get_column_letter(c0 + COLS_POR_MES)].width = 2

    for idx, m in enumerate(meses):
        _escribir_mes_trimestre(ws, m, anyo, datos, _col_inicio(idx), "COMPRAS", border)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    nombre_trim = TRIMESTRES[trimestre].replace(" ", "_")
    filename = f"{nombre_trim}_{anyo}_Compras.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
